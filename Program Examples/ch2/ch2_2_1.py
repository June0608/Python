# ch2_3.py
import openpyxl

wb = openpyxl.Workbook()                        # 建立空白的活頁簿
print("所有工作表名稱 = ", wb.sheetnames)       # 列印所有工作表
wb.create_sheet()                               # 建立新工作表
print("所有工作表名稱 = ", wb.sheetnames)       # 列印所有工作表
wb.create_sheet('First sheet', 0)               # 第 1 個工作表
print("所有工作表名稱 = ", wb.sheetnames)       # 列印所有工作表
wb.create_sheet('Third sheet', 2)               # 第 3 個工作表
print("所有工作表名稱 = ", wb.sheetnames)       # 列印所有工作表
wb.create_sheet('Fourth sheet', -1)             # 第 4 個工作表
print("所有工作表名稱 = ", wb.sheetnames)       # 列印所有工作表
wb.save('out2_3.xlsx')                          # 將活頁簿儲存





