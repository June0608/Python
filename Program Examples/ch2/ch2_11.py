# ch2_11.py
import openpyxl

fn = "data2_11.xlsx"
wb = openpyxl.load_workbook(fn)             # 開啟活頁簿
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
for sheet in wb.sheetnames:
    if sheet != "1月":
        ws = wb[sheet]
        wb.remove(ws)
src = wb.active
for i in range(2,13):
    dst = wb.copy_worksheet(src)
    month = str(i) + "月"
    dst.title = month
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
wb.save('out2_11.xlsx')                     # 將活頁簿儲存





