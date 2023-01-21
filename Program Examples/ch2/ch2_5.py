# ch2_5.py
import openpyxl

fn = "data2_5.xlsx"
wb = openpyxl.load_workbook(fn)             # 開啟活頁簿
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
src = wb.active
dst2 = wb.copy_worksheet(src)
dst2.title = "2025Q2"
dst3 = wb.copy_worksheet(src)
dst3.title = "2025Q3"
dst4 = wb.copy_worksheet(src)
dst4.title = "2025Q4"
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
wb.save('out2_5.xlsx')                      # 將活頁簿儲存





