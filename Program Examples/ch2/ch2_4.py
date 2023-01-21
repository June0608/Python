# ch2_4.py
import openpyxl

fn = "data2_4.xlsx"
wb = openpyxl.load_workbook(fn)             # 開啟活頁簿
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
src = wb.active
dst = wb.copy_worksheet(src)
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
wb.save('out2_4.xlsx')                      # 將活頁簿儲存





