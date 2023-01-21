# ch2_6_1.py
import openpyxl

fn = "data2_6_1.xlsx"
wb = openpyxl.load_workbook(fn)             # 開啟活頁簿
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
del wb['2025Q3']                        
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
wb.save('out2_6_1.xlsx')                    # 將活頁簿儲存





