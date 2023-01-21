# ch2_5_1.py
import openpyxl

fn = "data2_5_1.xlsx"
wb = openpyxl.load_workbook(fn)             # 開啟活頁簿
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
for sheet in wb.sheetnames:
    ws = wb[sheet]
    ws.title = sheet.replace('2025', '2026')
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
wb.save('out2_5_1.xlsx')                    # 將活頁簿儲存





