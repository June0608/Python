# ch2_10.py
import openpyxl

fn = "data2_10.xlsx"
wb = openpyxl.load_workbook(fn)             # 開啟活頁簿
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
src = wb.active
for i in range(2,13):
    dst = wb.copy_worksheet(src)
    month = str(i) + "月"
    dst.title = month
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
wb.save('out2_10.xlsx')                     # 將活頁簿儲存





