# ch2_6.py
import openpyxl

fn = "data2_6.xlsx"
wb = openpyxl.load_workbook(fn)             # 開啟活頁簿
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
sheet = wb['2025Q3']
wb.remove(sheet)
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
sheet = wb.worksheets[1]
wb.remove(sheet)
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
wb.save('out2_6.xlsx')                      # 將活頁簿儲存





