# ch2_8.py
import openpyxl

fn = "data2_8.xlsx"
wb = openpyxl.load_workbook(fn)             # 開啟活頁簿
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
ws = wb['2025Q3']
ws.sheet_state = "hidden"
print("所有工作表名稱 = ", wb.sheetnames)   # 列印所有工作表
wb.save('out2_8.xlsx')                      # 將活頁簿儲存





